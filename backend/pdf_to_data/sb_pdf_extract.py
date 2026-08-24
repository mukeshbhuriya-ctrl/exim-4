import json
import re
from datetime import datetime
from pathlib import Path

import pdfplumber
from openpyxl import Workbook, load_workbook


ROOT_FOLDER = Path(r"C:\Users\Khushal Makani\Documents\CODE\GFL\SB PDF OCR Extract")
INPUT_DIR = ROOT_FOLDER / "IN"
OUTPUT_DIR = ROOT_FOLDER / "OUT"
STATUS_DIR = ROOT_FOLDER / "STATUS"
TRACKER_PATH = STATUS_DIR / "PDF_Process_Tracker.xlsx"

BASE_COLUMNS = ["source_pdf", "Port Code", "SB No", "SB Date"]

PART3_COLUMNS = [
    "1INVSN",
    "2ITEMSN",
    "3.HS CD",
    "4.DESCRIPTION",
    "5.QUANTITY",
    "6UQC",
    "7.RATE",
    "8VALUE(F/C)",
    "9.FOB (INR)",
    "10.PMV",
    "11.DUTYAMT",
    "12.CESS RT",
    "13CESAMT",
    "14.DBKCLMD",
    "15.IGSTSTAT",
    "16. IGST VALUE",
    "17. IGST AMOUNT",
    "18SCHCOD",
    "19. SCHEME DESCRIPTION",
    "20. SQC MSR",
    "21. SQC UQC",
    "22. STATE OF ORIGIN",
    "23. DISTRICT OF ORIGIN",
    "24. PT Abroad",
    "25.COMP CESS",
    "26.END USE",
    "27.FTA BENEFIT AVAILED",
    "28. REWARD BENEFIT",
    "29. THIRD PARTY ITEM",
]

PART4_A_COLUMNS = [
    "1.INV SNO",
    "2.ITEM SNO",
    "3.DBK SNO.",
    "4.QTY/WT",
    "5.VALUE",
    "6.RATE",
    "7.DBK AMT",
    "8.STALEV",
    "9.CENLEV",
    "10.ROSCTL AMT",
]

PART4_H_COLUMNS = [
    "1.SNO",
    "2.INVOICENO",
    "3.INVOICEAMOUNT",
    "4.CURRENCY",
]

PART4_M_COLUMNS = [
    "1.INVSN",
    "2.ITMSN",
    "3.QUANTITY",
    "4.UQC",
    "5.NO.OF UNITS",
    "6. VALUE",
]

SHEET_CONFIGS = [
    ("Part3_Item_Details", PART3_COLUMNS, "part3_item_details"),
    ("Part4_A_Drawback", PART4_A_COLUMNS, "part4_a_drawback_rosl_claim"),
    ("Part4_H_Invoice", PART4_H_COLUMNS, "part4_h_invoice_details"),
    ("Part4_M_RODTEP", PART4_M_COLUMNS, "part4_m_rodtep_details"),
]

TRACKER_COLUMNS = [
    "document_key",
    "pdf_name",
    "status",
    "start_time",
    "end_time",
    "taken_time_seconds",
    "taken_time_hhmmss",
    "remarks",
    "excel_output_path",
    "json_output_path",
]


def clean_text(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def get_clean_lines(page_text):
    return [clean_text(line) for line in page_text.splitlines() if clean_text(line)]


def empty_row(columns):
    return {column: "" for column in columns}


def find_header_values(document_text):
    # Backward-compatible wrapper kept for older call-sites.
    return find_header_values_from_pages([document_text], "")


def extract_header_from_page_text(page_text):
    compact = " ".join(page_text.split())
    pattern = (
        r"Port\s*Code\s*SB\s*No\s*SB\s*Date\s*"
        r"(?:INDIAN\s+CUSTOMS\s+EDI\s+SYSTEM\s*)?"
        r"([A-Z0-9]{6})\s+(\d+)\s+([0-9]{2}-[A-Z]{3}-[0-9]{2})"
    )
    match = re.search(pattern, compact, flags=re.IGNORECASE)
    if not match:
        return None
    return (
        clean_text(match.group(1)).upper(),
        clean_text(match.group(2)),
        clean_text(match.group(3)).upper(),
    )


def extract_header_from_filename(pdf_stem):
    # Example stem: 927354001022026INHZA1SB22020220261809
    match = re.search(r"([A-Z]{6})SB", pdf_stem.upper())
    if not match:
        return None

    port_code = match.group(1)
    sb_match = re.match(r"(\d{7})", pdf_stem)
    sb_no = sb_match.group(1) if sb_match else ""
    return port_code, sb_no


def choose_most_common(values):
    if not values:
        return ""
    counts = {}
    for value in values:
        counts[value] = counts.get(value, 0) + 1
    return max(counts.items(), key=lambda item: item[1])[0]


def find_header_values_from_pages(page_texts, pdf_stem):
    page_headers = []
    for page_text in page_texts:
        parsed = extract_header_from_page_text(page_text)
        if parsed:
            page_headers.append(parsed)

    port_candidates = [item[0] for item in page_headers if item[0]]
    sb_candidates = [item[1] for item in page_headers if item[1]]
    date_candidates = [item[2] for item in page_headers if item[2]]

    port_code = choose_most_common(port_candidates)
    sb_no = choose_most_common(sb_candidates)
    sb_date = choose_most_common(date_candidates)

    if (not port_code or not sb_no) and pdf_stem:
        filename_header = extract_header_from_filename(pdf_stem)
        if filename_header:
            port_code = port_code or filename_header[0]
            sb_no = sb_no or filename_header[1]

    return {"Port Code": port_code, "SB No": sb_no, "SB Date": sb_date}


def is_noise_line(line):
    text = clean_text(line)
    if not text:
        return True
    if len(text) == 1 and text.isalpha() and text.upper() not in {"Y", "N"}:
        return True
    if re.fullmatch(r"[()/A-Z]+", text) and len(text) <= 4:
        return True
    return False


def find_first_meaningful_line(lines, start_idx, stop_markers):
    index = start_idx
    while index < len(lines):
        line = lines[index]
        if any(marker in line for marker in stop_markers):
            return "", index
        if not is_noise_line(line):
            return line, index + 1
        index += 1
    return "", index


def parse_part3_page(page_text):
    lines = get_clean_lines(page_text)
    rows = []
    index = 0

    while index < len(lines):
        if "1INVSN 2ITEMSN 3.HS CD" not in lines[index]:
            index += 1
            continue

        row = empty_row(PART3_COLUMNS)
        index += 1

        detail_lines = []
        while (
            index < len(lines)
            and "11.DUTYAMT" not in lines[index]
            and "1INVSN 2ITEMSN 3.HS CD" not in lines[index]
        ):
            if not is_noise_line(lines[index]):
                detail_lines.append(lines[index])
            index += 1

        primary_line = detail_lines[0] if detail_lines else ""
        extra_description = " ".join(detail_lines[1:]) if len(detail_lines) > 1 else ""

        detail_match = re.search(
            r"^(\d+)\s+(\d+)\s+(\d{6,10})(.*?)\s+([\d.,]+)\s+([A-Z]{2,5})\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)\s*$",
            primary_line,
        )
        if detail_match:
            row["1INVSN"] = clean_text(detail_match.group(1))
            row["2ITEMSN"] = clean_text(detail_match.group(2))
            row["3.HS CD"] = clean_text(detail_match.group(3))
            row["4.DESCRIPTION"] = clean_text(
                f"{detail_match.group(4)} {extra_description}".strip()
            )
            row["5.QUANTITY"] = clean_text(detail_match.group(5))
            row["6UQC"] = clean_text(detail_match.group(6))
            row["7.RATE"] = clean_text(detail_match.group(7))
            row["8VALUE(F/C)"] = clean_text(detail_match.group(8))
            row["9.FOB (INR)"] = clean_text(detail_match.group(9))
            row["10.PMV"] = clean_text(detail_match.group(10))
        else:
            row["4.DESCRIPTION"] = clean_text(" ".join(detail_lines))

        if index < len(lines) and "11.DUTYAMT" in lines[index]:
            index += 1
            values_11_18, index = find_first_meaningful_line(
                lines,
                index,
                stop_markers=["19. SCHEME DESCRIPTION", "1INVSN 2ITEMSN 3.HS CD"],
            )
            tokens = values_11_18.split()
            if tokens:
                # Typical OCR row: "Y LUT 19" maps to 14=Y, 15=LUT, 18=19.
                if len(tokens) >= 1:
                    row["14.DBKCLMD"] = tokens[0]
                if len(tokens) >= 2:
                    igst_status = tokens[1].upper()
                    if igst_status.endswith("LUT"):
                        igst_status = "LUT"
                    row["15.IGSTSTAT"] = igst_status
                if len(tokens) >= 3:
                    row["18SCHCOD"] = tokens[-1]

                middle_tokens = tokens[2:-1] if len(tokens) > 3 else []
                number_tokens = [
                    token for token in middle_tokens if re.fullmatch(r"[\d.,]+", token)
                ]
                if len(number_tokens) >= 1:
                    row["16. IGST VALUE"] = number_tokens[0]
                if len(number_tokens) >= 2:
                    row["17. IGST AMOUNT"] = number_tokens[1]

        if index < len(lines) and "19. SCHEME DESCRIPTION" in lines[index]:
            index += 1
            values_19_23, index = find_first_meaningful_line(
                lines,
                index,
                stop_markers=["24. PT Abroad", "1INVSN 2ITEMSN 3.HS CD"],
            )
            scheme_match = re.search(
                r"^(.*?)\s+([\d.,]+)\s+([A-Z]{2,5})\s+([A-Za-z ]+?)\s+([A-Za-z ]+)$",
                values_19_23,
            )
            if scheme_match:
                row["19. SCHEME DESCRIPTION"] = clean_text(scheme_match.group(1))
                row["20. SQC MSR"] = clean_text(scheme_match.group(2))
                row["21. SQC UQC"] = clean_text(scheme_match.group(3))
                row["22. STATE OF ORIGIN"] = clean_text(scheme_match.group(4))
                row["23. DISTRICT OF ORIGIN"] = clean_text(scheme_match.group(5))
            else:
                row["19. SCHEME DESCRIPTION"] = clean_text(values_19_23)

        # Move forward to the 24.* header, skipping OCR noise lines (e.g. single chars).
        while (
            index < len(lines)
            and "24. PT Abroad" not in lines[index]
            and "1INVSN 2ITEMSN 3.HS CD" not in lines[index]
        ):
            index += 1

        if index < len(lines) and "24. PT Abroad" in lines[index]:
            index += 1
            values_24_29, index = find_first_meaningful_line(
                lines,
                index,
                stop_markers=["1INVSN 2ITEMSN 3.HS CD", "GLOSSARY", "Scan QR Code"],
            )

            tokens = [
                token
                for token in values_24_29.split()
                if not (len(token) == 1 and token.isalpha() and token.upper() not in {"Y", "N"})
            ]
            if tokens:
                row["24. PT Abroad"] = tokens[0]
                token_index = 1

                if len(tokens) > 1:
                    comp_cess = tokens[1]
                    if len(tokens) > 2 and tokens[2] in {"INR", "USD", "EUR"}:
                        comp_cess = f"{comp_cess} {tokens[2]}"
                        token_index = 3
                    else:
                        token_index = 2
                    row["25.COMP CESS"] = comp_cess

                remaining = tokens[token_index:]
                if len(remaining) >= 4:
                    row["26.END USE"] = " ".join(remaining[:-3])
                    row["27.FTA BENEFIT AVAILED"] = remaining[-3]
                    row["28. REWARD BENEFIT"] = remaining[-2]
                    row["29. THIRD PARTY ITEM"] = remaining[-1]
                elif len(remaining) == 3:
                    row["26.END USE"] = remaining[0]
                    row["27.FTA BENEFIT AVAILED"] = remaining[1]
                    row["28. REWARD BENEFIT"] = remaining[2]
                elif len(remaining) == 2:
                    row["26.END USE"] = remaining[0]
                    row["27.FTA BENEFIT AVAILED"] = remaining[1]
                elif len(remaining) == 1:
                    row["26.END USE"] = remaining[0]

        rows.append(row)

    return rows


def parse_part4_a_page(page_text):
    lines = get_clean_lines(page_text)
    rows = []

    start_idx = -1
    for idx, line in enumerate(lines):
        if "1.INV SNO 2.ITEM SNO 3.DBK SNO." in line:
            start_idx = idx + 1
            break
    if start_idx == -1:
        return rows

    idx = start_idx
    while idx < len(lines):
        line = lines[idx]
        if line.startswith("B.") or "AA / DFIA" in line:
            break

        if re.match(r"^\d+\s+\d+", line):
            normalized_line = re.sub(r"([0-9])Y([0-9])", r"\1 \2", line)
            tokens = normalized_line.split()
            if len(tokens) >= 7:
                row = empty_row(PART4_A_COLUMNS)
                row["1.INV SNO"] = tokens[0]
                row["2.ITEM SNO"] = tokens[1]
                row["3.DBK SNO."] = tokens[2]
                row["4.QTY/WT"] = tokens[3]
                row["5.VALUE"] = tokens[4]
                row["6.RATE"] = tokens[5]
                row["7.DBK AMT"] = tokens[6]
                row["8.STALEV"] = tokens[7] if len(tokens) > 7 else ""
                row["9.CENLEV"] = tokens[8] if len(tokens) > 8 else ""
                row["10.ROSCTL AMT"] = tokens[9] if len(tokens) > 9 else ""
                rows.append(row)
        idx += 1

    return rows


def parse_part4_h_page(page_text):
    lines = get_clean_lines(page_text)
    rows = []

    header_idx = -1
    for idx, line in enumerate(lines):
        if (
            "1.SNO" in line
            and "2.INVOICE NO" in line
            and "3.INVOICE AMOUNT" in line
            and "4.CURRENCY" in line
        ):
            header_idx = idx
            break
    if header_idx == -1:
        return rows

    idx = header_idx + 1
    while idx < len(lines):
        line = lines[idx]
        if (
            line.startswith("I.CONTAINER")
            or line.startswith("Scan QR Code")
            or line.startswith("Page ")
            or line.startswith("NOITAMROFNI")
        ):
            break
        if re.match(r"^\d+\s+", line):
            tokens = line.split()
            if len(tokens) >= 4:
                row = empty_row(PART4_H_COLUMNS)
                row["1.SNO"] = tokens[0]
                row["2.INVOICENO"] = tokens[1]
                row["3.INVOICEAMOUNT"] = tokens[2]
                row["4.CURRENCY"] = tokens[3]
                rows.append(row)
        idx += 1

    return rows


def parse_part4_m_page(page_text):
    lines = get_clean_lines(page_text)
    rows = []

    start_idx = -1
    for idx, line in enumerate(lines):
        if "M. RODTEP DETAILS" in line:
            start_idx = idx
            break
    if start_idx == -1:
        return rows

    idx = start_idx + 1
    while idx < len(lines) and "1.INVSN2.ITMSN" not in lines[idx]:
        if lines[idx].startswith("N. REEXPORT"):
            return rows
        idx += 1
    if idx >= len(lines):
        return rows

    idx += 1
    while idx < len(lines):
        line = lines[idx]
        if line.startswith("N. REEXPORT") or line.startswith("Glossary") or line.startswith("Scan QR Code"):
            break
        if re.match(r"^\d+\s+\d+", line):
            tokens = line.split()
            if len(tokens) >= 6:
                row = empty_row(PART4_M_COLUMNS)
                row["1.INVSN"] = tokens[0]
                row["2.ITMSN"] = tokens[1]
                row["3.QUANTITY"] = tokens[2] if len(tokens) > 2 else ""
                row["4.UQC"] = tokens[3] if len(tokens) > 3 else ""
                row["5.NO.OF UNITS"] = tokens[4] if len(tokens) > 4 else ""
                row["6. VALUE"] = tokens[5] if len(tokens) > 5 else ""
                rows.append(row)
        idx += 1

    return rows


def attach_base_fields(rows, pdf_name, header_values, columns):
    if not rows:
        return []
    final_rows = []
    for row in rows:
        merged = empty_row(BASE_COLUMNS + columns)
        merged["source_pdf"] = pdf_name
        merged["Port Code"] = header_values.get("Port Code", "")
        merged["SB No"] = header_values.get("SB No", "")
        merged["SB Date"] = header_values.get("SB Date", "")
        for column in columns:
            merged[column] = clean_text(row.get(column, ""))
        final_rows.append(merged)
    return final_rows


def write_rows_to_sheet(sheet, rows, columns):
    sheet.append(columns)
    for row in rows:
        sheet.append([clean_text(row.get(column, "")) for column in columns])


def write_pdf_workbook(pdf_output_path, pdf_name, header_values, extracted):
    workbook = Workbook()
    first_sheet = True

    for sheet_name, section_columns, section_key in SHEET_CONFIGS:
        rows = attach_base_fields(
            extracted.get(section_key, []),
            pdf_name,
            header_values,
            section_columns,
        )
        columns = BASE_COLUMNS + section_columns

        if first_sheet:
            sheet = workbook.active
            sheet.title = sheet_name
            first_sheet = False
        else:
            sheet = workbook.create_sheet(sheet_name)

        write_rows_to_sheet(sheet, rows, columns)

    workbook.save(pdf_output_path)


def format_duration(seconds_value):
    total_seconds = max(0, int(round(seconds_value)))
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    seconds = total_seconds % 60
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"


def get_sheet_headers(sheet):
    if sheet.max_row < 1:
        return []
    return [clean_text(cell.value) for cell in sheet[1]]


def ensure_tracker_workbook(tracker_path):
    if tracker_path.exists():
        return
    tracker_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "PDF_Process_Tracker"
    sheet.append(TRACKER_COLUMNS)
    workbook.save(tracker_path)


def ensure_tracker_headers(sheet):
    if sheet.max_row == 1 and all(cell.value in (None, "") for cell in sheet[1]):
        for idx, header in enumerate(TRACKER_COLUMNS, start=1):
            sheet.cell(row=1, column=idx, value=header)
        return TRACKER_COLUMNS

    headers = get_sheet_headers(sheet)
    if not headers:
        for idx, header in enumerate(TRACKER_COLUMNS, start=1):
            sheet.cell(row=1, column=idx, value=header)
        return TRACKER_COLUMNS

    missing = [header for header in TRACKER_COLUMNS if header not in headers]
    if missing:
        start_col = len(headers) + 1
        for offset, header in enumerate(missing):
            sheet.cell(row=1, column=start_col + offset, value=header)
        headers.extend(missing)
    return headers


def get_tracker_context(tracker_path):
    ensure_tracker_workbook(tracker_path)
    workbook = load_workbook(tracker_path)
    sheet = workbook.active
    headers = ensure_tracker_headers(sheet)
    column_map = {header: idx + 1 for idx, header in enumerate(headers) if header}
    return workbook, sheet, column_map


def tracker_set_values(sheet, column_map, row_index, values):
    for key, value in values.items():
        col_index = column_map.get(key)
        if col_index:
            sheet.cell(row=row_index, column=col_index, value=value)


def tracker_get_value(sheet, column_map, row_index, key):
    col_index = column_map.get(key)
    if not col_index:
        return ""
    value = sheet.cell(row=row_index, column=col_index).value
    return clean_text(value)


def build_tracker_index(sheet, column_map):
    index = {}
    key_col = column_map.get("document_key")
    if not key_col:
        return index
    for row_index in range(2, sheet.max_row + 1):
        key_value = clean_text(sheet.cell(row=row_index, column=key_col).value)
        if key_value:
            index[key_value] = row_index
    return index


def upsert_pdf_entries_in_tracker(sheet, column_map, tracker_index, pdf_files):
    new_count = 0
    for pdf_path in pdf_files:
        document_key = pdf_path.stem
        row_index = tracker_index.get(document_key)
        if row_index is None:
            row_index = sheet.max_row + 1
            tracker_index[document_key] = row_index
            tracker_set_values(
                sheet,
                column_map,
                row_index,
                {
                    "document_key": document_key,
                    "pdf_name": pdf_path.name,
                    "status": "NEW",
                    "remarks": "Discovered new PDF",
                    "start_time": "",
                    "end_time": "",
                    "taken_time_seconds": "",
                    "taken_time_hhmmss": "",
                    "excel_output_path": "",
                    "json_output_path": "",
                },
            )
            new_count += 1
            continue

        current_status = tracker_get_value(sheet, column_map, row_index, "status").upper()
        update_values = {"pdf_name": pdf_path.name}
        if current_status == "IN_PROGRESS":
            update_values["status"] = "PENDING"
            update_values["remarks"] = "Previous run interrupted; moved to PENDING"
        tracker_set_values(sheet, column_map, row_index, update_values)

    return new_count


def process_pdf(pdf_path):
    with pdfplumber.open(str(pdf_path)) as pdf:
        page_texts = [page.extract_text() or "" for page in pdf.pages]

    header_values = find_header_values_from_pages(page_texts, pdf_path.stem)

    part3_rows = []
    part4_a_rows = []
    part4_h_rows = []
    part4_m_rows = []

    for page_text in page_texts:
        if "PART - III - ITEM DETAILS" in page_text:
            part3_rows.extend(parse_part3_page(page_text))
        if "A. DRAWBACK & ROSL CLAIM" in page_text:
            part4_a_rows.extend(parse_part4_a_page(page_text))
        if "H.INVOICE DETAILS" in page_text:
            part4_h_rows.extend(parse_part4_h_page(page_text))
        if "M. RODTEP DETAILS" in page_text:
            part4_m_rows.extend(parse_part4_m_page(page_text))

    return {
        "header": header_values,
        "part3_item_details": part3_rows,
        "part4_a_drawback_rosl_claim": part4_a_rows,
        "part4_h_invoice_details": part4_h_rows,
        "part4_m_rodtep_details": part4_m_rows,
    }


def process_all(root_folder, limit=None):
    input_dir = root_folder / "IN"
    output_dir = root_folder / "OUT"
    status_dir = root_folder / "STATUS"
    tracker_path = status_dir / "PDF_Process_Tracker.xlsx"

    input_dir.mkdir(parents=True, exist_ok=True)
    output_dir.mkdir(parents=True, exist_ok=True)
    status_dir.mkdir(parents=True, exist_ok=True)

    json_dir = output_dir / "JSON"
    excel_dir = output_dir / "EXCEL"
    json_dir.mkdir(parents=True, exist_ok=True)
    excel_dir.mkdir(parents=True, exist_ok=True)

    pdf_files = sorted(input_dir.glob("*.pdf"))
    if limit is not None:
        pdf_files = pdf_files[:limit]

    tracker_workbook, tracker_sheet, tracker_columns = get_tracker_context(tracker_path)
    tracker_index = build_tracker_index(tracker_sheet, tracker_columns)
    new_count = upsert_pdf_entries_in_tracker(
        tracker_sheet, tracker_columns, tracker_index, pdf_files
    )
    tracker_workbook.save(tracker_path)

    if not pdf_files:
        print(f"No PDFs found in: {input_dir}")
        print(f"Tracker updated: {tracker_path}")
        return

    processing_queue = []
    skipped_count = 0
    for pdf_path in pdf_files:
        row_index = tracker_index.get(pdf_path.stem)
        status = tracker_get_value(tracker_sheet, tracker_columns, row_index, "status").upper()
        if status in {"NEW", "PENDING"}:
            processing_queue.append(pdf_path)
        else:
            skipped_count += 1

    if not processing_queue:
        print(
            "No NEW/PENDING PDFs to process. "
            "Set tracker status to PENDING if you want reprocessing."
        )
        print(f"Tracker updated: {tracker_path}")
        return

    success_count = 0
    failure_count = 0

    for index, pdf_path in enumerate(processing_queue, start=1):
        print(f"[{index}/{len(processing_queue)}] Processing: {pdf_path.name}")
        row_index = tracker_index[pdf_path.stem]
        start_time = datetime.now()
        status = "FAILED"
        remarks = ""
        excel_path = excel_dir / f"{pdf_path.stem}.xlsx"
        json_path = json_dir / f"{pdf_path.stem}.json"
        excel_written = False
        json_written = False

        tracker_set_values(
            tracker_sheet,
            tracker_columns,
            row_index,
            {
                "pdf_name": pdf_path.name,
                "status": "IN_PROGRESS",
                "start_time": start_time.strftime("%Y-%m-%d %H:%M:%S"),
                "end_time": "",
                "taken_time_seconds": "",
                "taken_time_hhmmss": "",
                "remarks": "Processing started",
                "excel_output_path": "",
                "json_output_path": "",
            },
        )
        tracker_workbook.save(tracker_path)

        try:
            extracted = process_pdf(pdf_path)
            write_pdf_workbook(
                pdf_output_path=excel_path,
                pdf_name=pdf_path.name,
                header_values=extracted["header"],
                extracted=extracted,
            )
            excel_written = True

            json_output = {
                "source_pdf": pdf_path.name,
                "Port Code": extracted["header"].get("Port Code", ""),
                "SB No": extracted["header"].get("SB No", ""),
                "SB Date": extracted["header"].get("SB Date", ""),
                "part3_item_details": extracted["part3_item_details"],
                "part4_a_drawback_rosl_claim": extracted["part4_a_drawback_rosl_claim"],
                "part4_h_invoice_details": extracted["part4_h_invoice_details"],
                "part4_m_rodtep_details": extracted["part4_m_rodtep_details"],
            }
            with json_path.open("w", encoding="utf-8") as handle:
                json.dump(json_output, handle, ensure_ascii=False, indent=2)
            json_written = True

            status = "SUCCESS"
            remarks = "Processed successfully"
            success_count += 1
        except Exception as exc:
            status = "FAILED"
            remarks = str(exc)
            failure_count += 1
            print(f"  FAILED: {exc}")
        finally:
            end_time = datetime.now()
            duration_seconds = (end_time - start_time).total_seconds()
            tracker_set_values(
                tracker_sheet,
                tracker_columns,
                row_index,
                {
                    "pdf_name": pdf_path.name,
                    "status": status,
                    "start_time": start_time.strftime("%Y-%m-%d %H:%M:%S"),
                    "end_time": end_time.strftime("%Y-%m-%d %H:%M:%S"),
                    "taken_time_seconds": round(duration_seconds, 2),
                    "taken_time_hhmmss": format_duration(duration_seconds),
                    "remarks": remarks,
                    "excel_output_path": str(excel_path) if excel_written else "",
                    "json_output_path": str(json_path) if json_written else "",
                },
            )
            tracker_workbook.save(tracker_path)

    print(
        f"\nCompleted. Total IN PDFs: {len(pdf_files)}, New discovered: {new_count}, "
        f"Skipped (not NEW/PENDING): {skipped_count}, Processed: {len(processing_queue)}, "
        f"Success: {success_count}, Failed: {failure_count}"
    )
    print(f"Root folder: {root_folder}")
    print(f"Input folder: {input_dir}")
    print(f"Output folder: {output_dir}")
    print(f"Excel folder: {excel_dir}")
    print(f"JSON folder: {json_dir}")
    print(f"Tracker: {tracker_path}")


def main():
    process_all(root_folder=ROOT_FOLDER)


if __name__ == "__main__":
    main()
